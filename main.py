import asyncio
import json
import logging
import os
from dataclasses import dataclass, asdict, field
from pathlib import Path
from typing import List, Optional
from datetime import datetime
import re
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
from telethon.sync import TelegramClient, errors
from telethon.tl import types
from telethon.tl.functions.contacts import ImportContactsRequest, DeleteContactsRequest
from telethon.tl.functions.users import GetFullUserRequest
import nest_asyncio
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

nest_asyncio.apply()

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s", handlers=[logging.FileHandler("telegram_checker.log"), logging.StreamHandler()])
logger = logging.getLogger(__name__)
CONFIG_FILE = Path("config.json")
PROFILE_PHOTOS_DIR = Path("profile_photos")
RESULTS_DIR = Path("results")

@dataclass
class TelegramUser:
    id: int
    username: Optional[str]
    first_name: Optional[str]
    last_name: Optional[str]
    phone: str
    premium: bool
    verified: bool
    fake: bool
    bot: bool
    last_seen: str
    profile_photos: List[str] = field(default_factory=list)
    bio: Optional[str] = None

    @classmethod
    async def from_user(cls, client: TelegramClient, user: types.User, phone: str = "") -> 'TelegramUser':
        try:
            full_user = await client(GetFullUserRequest(user))
            return cls(
                id=user.id,
                username=user.username,
                first_name=getattr(user, 'first_name', None) or "",
                last_name=getattr(user, 'last_name', None) or "",
                phone=phone,
                premium=getattr(user, 'premium', False),
                verified=getattr(user, 'verified', False),
                fake=getattr(user, 'fake', False),
                bot=getattr(user, 'bot', False),
                last_seen=get_user_status(user.status),
                profile_photos=[],
                bio=full_user.full_user.about
            )
        except Exception as e:
            logger.error(f"Ошибка создания TelegramUser: {str(e)}")
            return cls(
                id=user.id,
                username=getattr(user, 'username', None),
                first_name=getattr(user, 'first_name', None) or "",
                last_name=getattr(user, 'last_name', None) or "",
                phone=phone,
                premium=getattr(user, 'premium', False),
                verified=getattr(user, 'verified', False),
                fake=getattr(user, 'fake', False),
                bot=getattr(user, 'bot', False),
                last_seen=get_user_status(getattr(user, 'status', None)),
                profile_photos=[],
                bio=None
            )

def get_user_status(status: types.TypeUserStatus) -> str:
    if isinstance(status, types.UserStatusOnline):
        return "В сети"
    elif isinstance(status, types.UserStatusOffline):
        return f"Был(а) в сети: {status.was_online.strftime('%Y-%m-%d %H:%M:%S')}"
    elif isinstance(status, types.UserStatusRecently):
        return "Недавно был(а) в сети"
    elif isinstance(status, types.UserStatusLastWeek):
        return "Был(а) на этой неделе"
    elif isinstance(status, types.UserStatusLastMonth):
        return "Был(а) в этом месяце"
    return "Неизвестно"

def validate_phone_number(phone: str) -> str:
    phone = re.sub(r'[^\d+]', '', phone.strip())
    if not phone.startswith('+'): phone = '+' + phone
    if not re.match(r'^\+\d{10,15}$', phone): raise ValueError(f"Неверный формат номера телефона: {phone}")
    return phone

def validate_username(username: str) -> str:
    username = username.strip().lstrip('@')
    if not re.match(r'^[A-Za-z]\w{3,30}[A-Za-z0-9]$', username): raise ValueError(f"Неверный формат имени пользователя: {username}")
    return username

class TelegramChecker:
    def __init__(self):
        self.config = self.load_config()
        self.client = None
        self.phone_code_hash = None
        PROFILE_PHOTOS_DIR.mkdir(exist_ok=True)
        RESULTS_DIR.mkdir(exist_ok=True)

    def load_config(self) -> dict:
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                logger.error(f"Ошибка загрузки конфигурации: {e}")
                return {}
        return {}

    def save_config(self):
        with open(CONFIG_FILE, 'w', encoding='utf-8') as f:
            json.dump(self.config, f, indent=2, ensure_ascii=False)

    async def initialize(self, window):
        if not self.config.get('api_id') or not self.config.get('api_hash'):
            window.output.insert(tk.END, "Первоначальная настройка - введите учетные данные Telegram API\n")
            window.output.insert(tk.END, "Вы можете получить их на https://my.telegram.org/apps\n")
            window.output.see(tk.END)
            return False

        if not self.client:
            self.client = TelegramClient('telegram_checker_session', self.config['api_id'], self.config['api_hash'])
            await self.client.connect()

        if await self.client.is_user_authorized():
            window.initialize_button.config(state=tk.DISABLED)
            return True
            
        return False

    async def send_code_request(self, phone):
        try:
            result = await self.client.send_code_request(phone)
            self.phone_code_hash = result.phone_code_hash
            return True
        except Exception as e:
            logger.error(f"Ошибка отправки кода: {str(e)}")
            return False

    async def sign_in(self, window, phone, code):
        try:
            if not self.phone_code_hash:
                await self.send_code_request(phone)
            
            await self.client.sign_in(phone, code, phone_code_hash=self.phone_code_hash)
            window.output.insert(tk.END, "Успешная авторизация.\n")
            window.output.see(tk.END)
            window.initialize_button.config(state=tk.DISABLED)
            return True
        except errors.SessionPasswordNeededError:
            window.output.insert(tk.END, "Требуется пароль 2FA. Пожалуйста, введите его.\n")
            window.output.see(tk.END)
            return False
        except errors.PhoneCodeInvalidError:
            window.output.insert(tk.END, "Неверный код подтверждения. Попробуйте еще раз.\n")
            window.output.see(tk.END)
            return False
        except Exception as e:
            window.output.insert(tk.END, f"Ошибка при входе: {str(e)}\n")
            window.output.see(tk.END)
            return False

    async def sign_in_2fa(self, password):
        await self.client.sign_in(password=password)

    async def download_all_profile_photos(self, user: types.User, user_data: TelegramUser):
        try:
            photos = await self.client.get_profile_photos(user)
            if not photos: return
            
            identifier = user_data.username or user_data.phone or str(user_data.id)
            user_photos_dir = PROFILE_PHOTOS_DIR / identifier
            user_photos_dir.mkdir(exist_ok=True)
            
            user_data.profile_photos = []
            for i, photo in enumerate(photos):
                photo_path = user_photos_dir / f"photo_{i}.jpg"
                await self.client.download_media(photo, file=photo_path)
                user_data.profile_photos.append(str(photo_path))
        except Exception as e:
            logger.error(f"Ошибка загрузки фото профиля для {user.id}: {str(e)}")

    async def check_phone_number(self, phone: str) -> Optional[TelegramUser]:
        try:
            phone = validate_phone_number(phone)
            try:
                user = await self.client.get_entity(phone)
                telegram_user = await TelegramUser.from_user(self.client, user, phone)
                await self.download_all_profile_photos(user, telegram_user)
                return telegram_user
            except:
                contact = types.InputPhoneContact(client_id=0, phone=phone, first_name="Test", last_name="User")
                result = await self.client(ImportContactsRequest([contact]))
                
                if not result.users: return None
                
                user = result.users[0]
                try:
                    full_user = await self.client.get_entity(user.id)
                    await self.client(DeleteContactsRequest(id=[user.id]))
                    telegram_user = await TelegramUser.from_user(self.client, full_user, phone)
                    await self.download_all_profile_photos(full_user, telegram_user)
                    return telegram_user
                finally:
                    try:
                        await self.client(DeleteContactsRequest(id=[user.id]))
                    except:
                        pass
        except Exception as e:
            logger.error(f"Ошибка проверки {phone}: {str(e)}")
            return None

    async def check_username(self, username: str) -> Optional[TelegramUser]:
        try:
            username = validate_username(username)
            user = await self.client.get_entity(username)
            if not isinstance(user, types.User): return None
            telegram_user = await TelegramUser.from_user(self.client, user, "")
            await self.download_all_profile_photos(user, telegram_user)
            return telegram_user
        except ValueError as e:
            logger.error(f"Неверное имя пользователя {username}: {str(e)}")
            return None
        except errors.UsernameNotOccupiedError:
            logger.error(f"Имя пользователя {username} не найдено")
            return None
        except Exception as e:
            logger.error(f"Ошибка проверки имени пользователя {username}: {str(e)}")
            return None

    async def process_phones(self, phones: List[str], window) -> dict:
        results = {}
        total_phones = len(phones)
        window.output.insert(tk.END, f"\nОбработка {total_phones} номеров телефонов...\n")
        window.output.see(tk.END)
        
        for i, phone in enumerate(phones, 1):
            try:
                phone = phone.strip()
                if not phone: continue
                window.output.insert(tk.END, f"Проверка {phone} ({i}/{total_phones})\n")
                window.output.see(tk.END)
                user = await self.check_phone_number(phone)
                results[phone] = asdict(user) if user else {"error": "Аккаунт Telegram не найден"}
            except ValueError as e:
                results[phone] = {"error": str(e)}
            except Exception as e:
                results[phone] = {"error": f"Непредвиденная ошибка: {str(e)}"}
        return results

    async def process_usernames(self, usernames: List[str], window) -> dict:
        results = {}
        total_usernames = len(usernames)
        window.output.insert(tk.END, f"\nОбработка {total_usernames} имен пользователей...\n")
        window.output.see(tk.END)
        
        for i, username in enumerate(usernames, 1):
            try:
                username = username.strip()
                if not username: continue
                window.output.insert(tk.END, f"Проверка {username} ({i}/{total_usernames})\n")
                window.output.see(tk.END)
                user = await self.check_username(username)
                results[username] = asdict(user) if user else {"error": "Аккаунт Telegram не найден"}
            except ValueError as e:
                results[username] = {"error": str(e)}
            except Exception as e:
                results[username] = {"error": f"Непредвиденная ошибка: {str(e)}"}
        return results

    def export_to_excel(self, results: dict, filename: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Результаты проверки"

        headers = ["Идентификатор", "ID", "Имя пользователя", "Имя", "Фамилия", "Телефон", "Premium", "Верифицирован", 
                   "Фейк", "Бот", "Последняя активность", "Био", "Количество фото профиля"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

        for row, (identifier, data) in enumerate(results.items(), 2):
            if "error" in data:
                ws.cell(row=row, column=1, value=identifier)
                ws.cell(row=row, column=2, value=data["error"])
                continue

            ws.cell(row=row, column=1, value=identifier)
            ws.cell(row=row, column=2, value=data.get("id", ""))
            ws.cell(row=row, column=3, value=data.get("username", ""))
            ws.cell(row=row, column=4, value=data.get("first_name", ""))
            ws.cell(row=row, column=5,value=data.get("last_name", ""))
            ws.cell(row=row, column=6, value=data.get("phone", ""))
            ws.cell(row=row, column=7, value="Да" if data.get("premium", False) else "Нет")
            ws.cell(row=row, column=8, value="Да" if data.get("verified", False) else "Нет")
            ws.cell(row=row, column=9, value="Да" if data.get("fake", False) else "Нет")
            ws.cell(row=row, column=10, value="Да" if data.get("bot", False) else "Нет")
            ws.cell(row=row, column=11, value=data.get("last_seen", ""))
            ws.cell(row=row, column=12, value=data.get("bio", ""))
            ws.cell(row=row, column=13, value=len(data.get("profile_photos", [])))

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(filename)

class TelegramCheckerGUI:
    def __init__(self, master):
        self.master = master
        master.title("Telegram Account Checker")
        master.geometry("800x800")

        self.checker = TelegramChecker()
        self.loop = asyncio.get_event_loop()

        self.create_widgets()
        self.auto_initialize()

    def create_widgets(self):
        ttk.Label(self.master, text="Telegram Account Checker", font=("Helvetica", 16)).pack(pady=10)

        main_frame = ttk.Frame(self.master)
        main_frame.pack(padx=20, pady=5, fill=tk.BOTH, expand=True)

        input_frame = ttk.Frame(main_frame)
        input_frame.pack(fill=tk.X, padx=10, pady=5)

        input_frame.columnconfigure(1, weight=1)

        ttk.Label(input_frame, text="API ID:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.api_id = ttk.Entry(input_frame)
        self.api_id.grid(row=0, column=1, sticky=tk.EW, pady=5)

        ttk.Label(input_frame, text="API Hash:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.api_hash = ttk.Entry(input_frame)
        self.api_hash.grid(row=1, column=1, sticky=tk.EW, pady=5)

        ttk.Label(input_frame, text="Телефон:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.phone = ttk.Entry(input_frame)
        self.phone.grid(row=2, column=1, sticky=tk.EW, pady=5)

        ttk.Label(input_frame, text="Код подтверждения:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.code = ttk.Entry(input_frame)
        self.code.grid(row=3, column=1, sticky=tk.EW, pady=5)

        ttk.Label(input_frame, text="2FA пароль (если есть):").grid(row=4, column=0, sticky=tk.W, pady=5)
        self.twofa = ttk.Entry(input_frame, show="*")
        self.twofa.grid(row=4, column=1, sticky=tk.EW, pady=5)

        self.initialize_button = ttk.Button(input_frame, text="Инициализировать клиент", command=self.initialize_client)
        self.initialize_button.grid(row=5, column=0, columnspan=2, pady=10)

        ttk.Label(input_frame, text="Введите номера телефонов или имена пользователей (через запятую):").grid(row=6, column=0, columnspan=2, sticky=tk.W, pady=5)
        self.input = scrolledtext.ScrolledText(input_frame, height=5)
        self.input.grid(row=7, column=0, columnspan=2, sticky=tk.EW, pady=5)

        radio_frame = ttk.Frame(input_frame)
        radio_frame.grid(row=8, column=0, columnspan=2, sticky=tk.EW, pady=5)
        
        self.check_type = tk.StringVar(value="phone")
        ttk.Radiobutton(radio_frame, text="Проверить номера телефонов", variable=self.check_type, value="phone").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(radio_frame, text="Проверить имена пользователей", variable=self.check_type, value="username").pack(side=tk.LEFT, padx=5)

        button_frame = ttk.Frame(input_frame)
        button_frame.grid(row=9, column=0, columnspan=2, sticky=tk.EW, pady=10)
        
        ttk.Button(button_frame, text="Проверить", command=self.check).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Экспорт в Excel", command=self.export_to_excel).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Очистить учетные данные", command=self.clear_credentials).pack(side=tk.LEFT, padx=5)

        self.output = scrolledtext.ScrolledText(main_frame, height=10)
        self.output.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        if self.checker.config:
            self.api_id.insert(0, str(self.checker.config.get('api_id', '')))
            self.api_hash.insert(0, self.checker.config.get('api_hash', ''))
            self.phone.insert(0, self.checker.config.get('phone', ''))

    def auto_initialize(self):
        self.loop.run_until_complete(self.async_auto_initialize())

    async def async_auto_initialize(self):
        if self.checker.config.get('api_id') and self.checker.config.get('api_hash'):
            initialized = await self.checker.initialize(self)
            if initialized:
                self.output.insert(tk.END, "Клиент успешно инициализирован.\n")
                self.output.see(tk.END)
                self.initialize_button.config(state=tk.DISABLED)

    def initialize_client(self):
        self.loop.run_until_complete(self.async_initialize_client())

    async def async_initialize_client(self):
        try:
            if not self.checker.config.get('api_id') or not self.checker.config.get('api_hash'):
                self.checker.config['api_id'] = int(self.api_id.get())
                self.checker.config['api_hash'] = self.api_hash.get()
                self.checker.config['phone'] = validate_phone_number(self.phone.get())
                self.checker.save_config()

            initialized = await self.checker.initialize(self)
            
            if not initialized:
                phone = self.phone.get()
                code = self.code.get()
                
                if code:
                    signed_in = await self.checker.sign_in(self, phone, code)
                    if signed_in:
                        self.output.insert(tk.END, "Клиент успешно инициализирован.\n")
                        self.output.see(tk.END)
                        self.initialize_button.config(state=tk.DISABLED)
                    return
                else:
                    sent = await self.checker.send_code_request(phone)
                    if sent:
                        self.output.insert(tk.END, "Код подтверждения отправлен. Пожалуйста, введите его и нажмите 'Инициализировать клиент' снова.\n")
                        self.output.see(tk.END)
                    return

            self.output.insert(tk.END, "Клиент инициализирован успешно.\n")
            self.output.see(tk.END)
            
        except Exception as e:
            self.output.insert(tk.END, f"Ошибка: {str(e)}\n")
            self.output.see(tk.END)

    def check(self):
        self.loop.run_until_complete(self.async_check())

    async def async_check(self):
        if not self.checker.client:
            messagebox.showerror("Ошибка", "Сначала инициализируйте клиент")
            return

        if not await self.checker.client.is_user_authorized():
            code = self.code.get()
            if not code:
                messagebox.showerror("Ошибка", "Введите код подтверждения")
                return
            
            signed_in = await self.checker.sign_in(self, self.phone.get(), code)
            if not signed_in:
                twofa = self.twofa.get()
                if not twofa:
                    messagebox.showerror("Ошибка", "Введите пароль 2FA")
                    return
                await self.checker.sign_in_2fa(twofa)

        input_data = [x.strip() for x in self.input.get("1.0", tk.END).strip().split(',') if x.strip()]
        if not input_data:
            messagebox.showerror("Ошибка", "Введите данные для проверки")
            return

        if self.check_type.get() == "phone":
            results = await self.checker.process_phones(input_data, self)
        else:
            results = await self.checker.process_usernames(input_data, self)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        usernames_str = "_".join(input_data[:3])
        if len(input_data) > 3:
            usernames_str += f"_and_{len(input_data)-3}_more"
        output_file = RESULTS_DIR / f"check_{usernames_str}_{timestamp}.json"

        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(results, f, indent=2, ensure_ascii=False)

        self.output.insert(tk.END, f"\nРезультаты сохранены в {output_file}\n")
        self.output.insert(tk.END, "\nКраткие результаты:\n")

        for identifier, data in results.items():
            if "error" in data:
                self.output.insert(tk.END, f"❌ {identifier}: {data['error']}\n")
            else:
                status = f"✓ {identifier}: {data.get('first_name', '')} {data.get('last_name', '')} (@{data.get('username', 'нет имени пользователя')})"
                if data.get('profile_photos'):
                    status += f" - загружено {len(data['profile_photos'])} фото профиля"
                self.output.insert(tk.END, status + "\n")

        self.output.insert(tk.END, "\nПодробные результаты сохранены в JSON файле.\n")
        self.output.see(tk.END)

        self.last_results = results

    def export_to_excel(self):
        if not hasattr(self, 'last_results'):
            json_files = list(RESULTS_DIR.glob('*.json'))
            if not json_files:
                messagebox.showerror("Ошибка", "Нет данных для экспорта. Сначала выполните проверку.")
                return
            latest_file = max(json_files, key=os.path.getctime)
            with open(latest_file, 'r', encoding='utf-8') as f:
                self.last_results = json.load(f)

        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if not file_path:
            return

        try:
            self.checker.export_to_excel(self.last_results, file_path)
            messagebox.showinfo("Успех", f"Данные успешно экспортированы в {file_path}")
        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось экспортировать данные: {str(e)}")

    def clear_credentials(self):
        if messagebox.askyesno("Подтверждение", "Вы уверены, что хотите очистить учетные данные?"):
            if CONFIG_FILE.exists():
                CONFIG_FILE.unlink()
            if self.checker.client:
                self.checker.client.disconnect()
                self.checker.client = None
            session_file = Path('telegram_checker_session.session')
            if session_file.exists():
                try:
                    session_file.unlink()
                except PermissionError:
                    messagebox.showwarning("Предупреждение", "Не удалось удалить файл сессии. Пожалуйста, закройте все программы, использующие Telegram, и попробуйте снова.")
            messagebox.showinfo("Информация", "Учетные данные очищены. Пожалуйста, перезапустите программу.")
            self.master.quit()

def main():
    root = tk.Tk()
    app = TelegramCheckerGUI(root)
    root.mainloop()

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nПрограмма прервана пользователем")
    except Exception as e:
        print(f"\nПроизошла ошибка: {str(e)}")
        logger.exception("Необработанное исключение")
